import io
import requests
import pandas as pd
import openpyxl
import os
import re

MESES = ['AGOSTO', 'SETEMBRO', 'OUTUBRO', 'NOVEMBRO', 'DEZEMBRO', 'JANEIRO']
LOCAL_BACKUP_PATH = os.path.join(os.path.dirname(__file__), "Encontro_de_Contas_Backup.xlsx")

def extract_file_or_sheet_id(url_or_id):
    if not url_or_id:
        return None
    # Check for Google Sheets /spreadsheets/d/ID
    match_sheets = re.search(r'/spreadsheets/d/([a-zA-Z0-9-_]+)', url_or_id)
    if match_sheets:
        return match_sheets.group(1), 'sheets'
    # Check for Google Drive /file/d/ID or ?id=ID
    match_drive = re.search(r'/file/d/([a-zA-Z0-9-_]+)', url_or_id)
    if match_drive:
        return match_drive.group(1), 'drive'
    match_id_param = re.search(r'[?&]id=([a-zA-Z0-9-_]+)', url_or_id)
    if match_id_param:
        return match_id_param.group(1), 'drive'
    # If raw ID passed
    if re.match(r'^[a-zA-Z0-9-_]{20,}$', url_or_id):
        return url_or_id, 'drive'
    return None, None

def load_data_from_google_drive(url_or_id):
    """
    Downloads and parses the spreadsheet directly from Google Drive / Google Sheets.
    Returns: dict of DataFrames for each month + boolean success status + message.
    """
    doc_id, doc_type = extract_file_or_sheet_id(url_or_id)
    if not doc_id:
        return load_data_from_local(), False, "ID ou Link inválido do Google Drive."

    try:
        session = requests.Session()
        if doc_type == 'sheets':
            download_url = f"https://docs.google.com/spreadsheets/d/{doc_id}/export?format=xlsx"
        else:
            download_url = f"https://docs.google.com/uc?export=download&id={doc_id}&confirm=t"
            
        resp = session.get(download_url, timeout=12)
        if resp.status_code != 200 or len(resp.content) < 1000:
            return load_data_from_local(), False, f"Não foi possível baixar (Status {resp.status_code}). Verifique se o compartilhamento está como 'Qualquer pessoa com o link'."
            
        # Parse Excel bytes
        excel_bytes = io.BytesIO(resp.content)
        wb = openpyxl.load_workbook(excel_bytes, data_only=True)
        
        # Save local copy for offline backup
        try:
            with open(LOCAL_BACKUP_PATH, "wb") as f:
                f.write(resp.content)
        except Exception:
            pass

        return parse_workbook_data(wb), True, "Sincronizado com sucesso com o Google Drive!"
    except Exception as e:
        # Fallback to local
        data = load_data_from_local()
        return data, False, f"Aviso de sincronização: {str(e)}. Usando dados locais salvos."

def load_data_from_local():
    """Loads fallback data from local database or file"""
    # Check if backup xlsx exists
    if os.path.exists(LOCAL_BACKUP_PATH):
        try:
            wb = openpyxl.load_workbook(LOCAL_BACKUP_PATH, data_only=True)
            return parse_workbook_data(wb)
        except Exception:
            pass
            
    # Default fallback data if empty
    import database as db
    db.init_db()
    all_notas = db.get_all_notas()
    
    result = {}
    for m in MESES:
        df_m = all_notas[all_notas['mes'] == m] if not all_notas.empty else pd.DataFrame()
        result[m] = df_m
    return result

def parse_workbook_data(wb):
    result = {}
    for m in MESES:
        if m in wb.sheetnames:
            ws = wb[m]
            rows = []
            for r in range(8, 32):
                d = ws.cell(row=r, column=1).value
                nf = ws.cell(row=r, column=2).value
                val = ws.cell(row=r, column=3).value
                
                # Check for hyperlink in Col D
                cell_d = ws.cell(row=r, column=4)
                link = None
                if cell_d.hyperlink:
                    link = cell_d.hyperlink.target
                elif cell_d.value and isinstance(cell_d.value, str):
                    if "http" in cell_d.value:
                        link = cell_d.value
                
                obs = ws.cell(row=r, column=5).value
                
                if nf is not None or val is not None:
                    # Format date
                    if hasattr(d, 'strftime'):
                        d_str = d.strftime('%Y-%m-%d')
                    else:
                        d_str = str(d) if d else ''
                        
                    try:
                        val_float = float(val) if val is not None else 0.0
                    except (ValueError, TypeError):
                        val_float = 0.0
                        
                    rows.append({
                        'id': r,
                        'mes': m,
                        'data': d_str,
                        'numero_nf': str(nf) if nf is not None else '',
                        'valor': val_float,
                        'link_drive': link or '',
                        'observacao': str(obs) if obs is not None else ''
                    })
            result[m] = pd.DataFrame(rows) if rows else pd.DataFrame(columns=['id', 'mes', 'data', 'numero_nf', 'valor', 'link_drive', 'observacao'])
        else:
            result[m] = pd.DataFrame(columns=['id', 'mes', 'data', 'numero_nf', 'valor', 'link_drive', 'observacao'])
    return result
